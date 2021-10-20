import os
import json
import openpyxl as op
from datetime import datetime, timedelta, date
from pathlib import Path

src_path = Path(__file__).parent
main_path = src_path.parents[1]
data_path = src_path.parent / 'data'
correo_path = src_path.parent / 'email bot'
excel_path = src_path.parent / 'excel'


def verificacion_carpetas():
    if os.path.exists(data_path) and os.path.exists(correo_path) and os.path.exists(excel_path):
        print("Carpetas esenciales existen")
        True
    else:
        print("Creacion de carpetas esenciales")
        if not os.path.exists(data_path):
            os.mkdir(f'{data_path}')
        if not os.path.exists(excel_path):
            os.mkdir(f'{excel_path}')
        print("Carpetas creadas")

def verificacion_archivo():
    if os.path.exists(data_path / 'correo.json') and os.path.exists(data_path / 'Resale certificate Camlem 2021.pdf'):
        print("Archivos 'correo.json' y 'Resale certificate Camlem 2021.pdf' se han encontrado")
        return True
    else:
        raise Exception(f"Archivos 'correo.json' o 'Resale certificate Camlem 2021.pdf' no encontrados en la carpeta {data_path}")

def archivo_fijo(): 
    ruta = data_path / 'Resale certificate Camlem 2021.pdf'
    return ruta


def read_json():
    with open(data_path / 'correo.json', encoding='utf-8') as archivo_json:
        return json.load(archivo_json)

def download_path():
    for root, _, _ in os.walk(main_path):
        if('download' in root.lower()):
            return root

def obtencion_archivos(nombre_carpeta):
    listado_carpeta = os.listdir(download_path())
    # Declaro una variable para guardar cada una de las carpetas que cumplen las condiciones de fecha
    carpetas = {}
    for carpeta in listado_carpeta:
        if nombre_carpeta in carpeta:
            carpetas[f'{download_path()}\\{carpeta}'] = os.listdir(f'{download_path()}\\{carpeta}')
            return carpetas

def listar_fechas(fecha_inicio, fecha_fin):
    fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin = datetime.strptime(fecha_fin,"%Y-%m-%d")

    if fecha_inicio < fecha_fin:
        lista_fechas = [(fecha_inicio + timedelta(days=d)).strftime("%Y-%m-%d") for d in range((fecha_fin - fecha_inicio).days + 1)]
        return lista_fechas
    elif fecha_fin == fecha_inicio:
        lista_fechas = [fecha_inicio]
        return lista_fechas
    else: 
        print("La fecha de inicio es una fecha mayor a la fecha final. Ingrese un intervalo de fechas correcta.")
        return False

def eliminar_log_error():
    if os.path.exists(f'{src_path.parent}\\log.txt'):
        os.remove(f'{src_path.parent}\\log.txt')

def log_error(carpeta, archivo):
    if os.path.exists(f'{src_path.parent}\\log.txt') == False:
        with open(f'{src_path.parent}\\log.txt', 'w') as error:
            error.write(f'*{datetime.today()} ---- Ha ocurrido un error con el archivo {archivo} en la carpeta {carpeta}' + "\n")
    else: 
        with open(f'{src_path.parent}\\log.txt', 'a') as error:
            error.write(f'*{datetime.today()} ---- Ha ocurrido un error con el archivo {archivo} en la carpeta {carpeta}' + "\n")

def crear_archivo_excel(nombre_archivo):
    if os.path.exists(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx'):
        archivo = op.load_workbook(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx')
        return archivo
    else: 
        archivo = op.Workbook()
        archivo.worksheets[0].title = "Informe"
        hoja = archivo.get_sheet_by_name("Informe")
        # Tamaños predeterminados dados
        hoja.column_dimensions['A'].width = 25
        # Coloco titulo 
        hoja.cell(row = 1, column = 1).value = "Nº ORDEN"
        return archivo

def escribir_excel(archivo_a_modificar, orden):
    #Busco la hoja donde tengo que modificar (General en cada archivo creado)
    hoja_a_modificar = archivo_a_modificar.active
    #Selecciono la ultima fila 
    ultima_fila = hoja_a_modificar.max_row
    #Escribo en la columna uno que es el numero de orden:
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).value = orden

def guardar_archivo_excel(archivo, nombre_archivo):
    # Guardo 
    archivo.save(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx')
    # Finalizo la sesion del archivo
    archivo.close()